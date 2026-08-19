import os
import json
import datetime
import base64
import sys
import calendar
import re
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure, DuplicateKeyError
from functools import wraps
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openai

app = Flask(__name__, static_folder='public', static_url_path='')
CORS(app)

MONGO_URI = os.environ.get('MONGO_URI')
DB_NAME = os.environ.get('DB_NAME', 'silver_sands_hrms')
AUTO_MIGRATE = os.environ.get('AUTO_MIGRATE', 'false').lower() == 'true'
FLASK_DEBUG = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY')

# Configure OpenAI - Works with both old and new versions
openai_client = None
OPENAI_AVAILABLE = False

if OPENAI_API_KEY:
    try:
        # Try new version first
        try:
            from openai import OpenAI
            openai_client = OpenAI(api_key=OPENAI_API_KEY)
            OPENAI_AVAILABLE = True
            print(" OpenAI configured successfully (new version)")
        except (ImportError, TypeError) as e:
            print(f" New OpenAI version not available, trying old version: {e}")
            # Try old version
            openai.api_key = OPENAI_API_KEY
            openai_client = openai
            OPENAI_AVAILABLE = True
            print(" OpenAI configured successfully (old version)")
    except Exception as e:
        print(f" OpenAI configuration error: {e}")
        OPENAI_AVAILABLE = False
else:
    print(" OPENAI_API_KEY not found in environment variables")

print("\n" + "=" * 55)
print("    SILVER SANDS SALIMA HRMS - STARTING")
print("=" * 55)
print(f"Database Name: {DB_NAME}")
print(f"Auto-Migration: {AUTO_MIGRATE}")
print(f"Debug Mode: {FLASK_DEBUG}")
print(f"OpenAI: {'Enabled' if OPENAI_AVAILABLE else 'Disabled'}")
print("=" * 55)

client = None
db = None

try:
    if not MONGO_URI:
        print(" MONGO_URI environment variable is not set!")
        sys.exit(1)
    
    masked_uri = MONGO_URI[:30] + "..." if len(MONGO_URI) > 30 else MONGO_URI
    print(f" Connecting to MongoDB: {masked_uri}")
    
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=10000)
    client.admin.command('ping')
    print(" MongoDB ping successful!")
    
    db = client[DB_NAME]
    print(f" Connected to MongoDB Atlas: {DB_NAME}")
    
    collections = db.list_collection_names()
    print(f" Available collections: {collections if collections else '(none yet)'}")
    
    try:
        attendance_collection = db['attendance']
        attendance_collection.create_index([('period', 1), ('employee_id', 1)])
        attendance_collection.create_index([('period', 1), ('date', 1)])
        
        sessions_collection = db['sessions']
        sessions_collection.create_index([('user_id', 1)])
        print(" Created MongoDB indexes")
    except Exception as e:
        print(f" Index creation skipped: {e}")
    
except Exception as e:
    print(f" MongoDB connection failed: {e}")
    sys.exit(1)

users_collection = db['users']
employees_collection = db['employees']
attendance_collection = db['attendance']
activity_log_collection = db['activity_logs']
sessions_collection = db['sessions']

def create_session(user_id, token):
    try:
        sessions_collection.delete_many({'user_id': user_id})
        
        session = {
            'user_id': user_id,
            'token': token,
            'created_at': datetime.datetime.now().isoformat(),
            'expires_at': (datetime.datetime.now() + datetime.timedelta(days=1)).isoformat()
        }
        sessions_collection.insert_one(session)
        return True
    except Exception as e:
        print(f"Error creating session: {e}")
        return False

def validate_session(token, user_id):
    try:
        session = sessions_collection.find_one({
            'token': token,
            'user_id': user_id
        })
        if not session:
            return False
        
        expires_at = datetime.datetime.fromisoformat(session['expires_at'])
        if expires_at < datetime.datetime.now():
            sessions_collection.delete_one({'_id': session['_id']})
            return False
        
        return True
    except Exception as e:
        print(f"Error validating session: {e}")
        return False

def clear_user_sessions(user_id):
    try:
        sessions_collection.delete_many({'user_id': user_id})
        return True
    except Exception as e:
        print(f"Error clearing sessions: {e}")
        return False

def log_activity(user, action, details=None, success=True, ip_address=None):
    try:
        log_entry = {
            "user_id": user.get('_id') if user else 'system',
            "username": user.get('username') if user else 'system',
            "user_role": user.get('role') if user else 'system',
            "action": action,
            "details": details or {},
            "ip_address": ip_address or request.remote_addr if request else 'unknown',
            "success": success,
            "timestamp": datetime.datetime.now().isoformat()
        }
        activity_log_collection.insert_one(log_entry)
        return True
    except Exception as e:
        print(f"Error logging activity: {e}")
        return False

def get_activity_logs(user, filters=None, limit=500, skip=0):
    try:
        query = filters or {}
        user_role = user.get('role') if user else None
        user_id = user.get('_id') if user else None
        
        if user_role != 'IT Specialist' and user_id:
            query['user_id'] = user_id
        
        logs = list(activity_log_collection.find(query)
                    .sort('timestamp', -1)
                    .skip(skip)
                    .limit(limit))
        
        for log in logs:
            log['_id'] = str(log['_id'])
        
        return logs
    except Exception as e:
        print(f"Error retrieving logs: {e}")
        return []

def get_activity_logs_count(user, filters=None):
    try:
        query = filters or {}
        user_role = user.get('role') if user else None
        user_id = user.get('_id') if user else None
        
        if user_role != 'IT Specialist' and user_id:
            query['user_id'] = user_id
        
        return activity_log_collection.count_documents(query)
    except Exception as e:
        print(f"Error counting logs: {e}")
        return 0

def auto_mark_day_offs(period):
    try:
        print(" Auto-marking day-offs for period: " + period)
        
        employees = list(employees_collection.find({'day_off': {'$ne': ''}}, {'_id': 1, 'day_off': 1}))
        
        if not employees:
            print("   No employees with scheduled day-offs")
            return 0
        
        year, month = map(int, period.split('-'))
        prev_month = 12 if month == 1 else month - 1
        prev_year = year - 1 if month == 1 else year
        
        period_dates = []
        
        days_in_prev = calendar.monthrange(prev_year, prev_month)[1]
        for d in range(27, days_in_prev + 1):
            date_str = f"{prev_year}-{str(prev_month).zfill(2)}-{str(d).zfill(2)}"
            period_dates.append(date_str)
        
        days_in_curr = calendar.monthrange(year, month)[1]
        for d in range(1, min(27, days_in_curr + 1)):
            date_str = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"
            period_dates.append(date_str)
        
        day_map = {
            'Monday': 0, 'Tuesday': 1, 'Wednesday': 2,
            'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6
        }
        
        bulk_operations = []
        
        for emp in employees:
            emp_id = emp['_id']
            day_off = emp.get('day_off', '')
            
            if not day_off:
                continue
            
            target_day = day_map.get(day_off)
            if target_day is None:
                continue
            
            for date_str in period_dates:
                date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                if date_obj.weekday() == target_day:
                    existing = attendance_collection.find_one({
                        'period': period,
                        'employee_id': emp_id,
                        'date': date_str
                    }, {'_id': 1})
                    
                    if not existing:
                        bulk_operations.append({
                            "period": period,
                            "employee_id": emp_id,
                            "date": date_str,
                            "status": "O",
                            "created_at": datetime.datetime.now().isoformat(),
                            "updated_at": datetime.datetime.now().isoformat(),
                            "auto_marked": True
                        })
        
        if bulk_operations:
            attendance_collection.insert_many(bulk_operations)
            print(f" Auto-marked {len(bulk_operations)} day-off records")
            return len(bulk_operations)
        
        print("   No new day-offs to mark")
        return 0
        
    except Exception as e:
        print(f" Error auto-marking day-offs: {e}")
        return 0

def auto_migrate_data():
    DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
    
    if not os.path.exists(DATA_DIR):
        print("No data directory found - skipping migration")
        return
    
    print("\n" + "=" * 55)
    print("    AUTO-MIGRATION STARTED")
    print("=" * 55)
    print(f"Data directory: {DATA_DIR}")
    
    users_file = os.path.join(DATA_DIR, 'users.json')
    if os.path.exists(users_file):
        print("\nMigrating users...")
        try:
            with open(users_file, 'r') as f:
                users_data = json.load(f)
            
            migrated = 0
            for username, user_data in users_data.items():
                try:
                    users_collection.insert_one({
                        "_id": user_data.get('id', username),
                        "username": username,
                        "password": user_data['password'],
                        "name": user_data['name'],
                        "role": user_data.get('role', 'User'),
                        "department": user_data.get('department', ''),
                        "created_at": user_data.get('createdAt', datetime.datetime.now().isoformat())
                    })
                    migrated += 1
                    print(f"   User: {username}")
                except DuplicateKeyError:
                    print(f"   User already exists: {username}")
            print(f"   Migrated {migrated} users")
        except Exception as e:
            print(f"   Error migrating users: {e}")
    else:
        print("\nNo users.json found - skipping")
    
    employees_file = os.path.join(DATA_DIR, 'employees.json')
    if os.path.exists(employees_file):
        print("\nMigrating employees...")
        try:
            with open(employees_file, 'r') as f:
                employees_data = json.load(f)
            
            migrated = 0
            for emp in employees_data:
                try:
                    existing = employees_collection.find_one({'_id': emp['id']})
                    if existing:
                        print(f"   Employee already exists: {emp['name']}")
                        continue
                    
                    employee_no = emp.get('employeeNo', '')
                    if not employee_no or employee_no == '':
                        employee_no = generate_employee_number(emp.get('department', ''))
                    
                    day_off = emp.get('day_off', '')
                    
                    employees_collection.insert_one({
                        "_id": emp['id'],
                        "employee_no": employee_no,
                        "name": emp['name'],
                        "department": emp['department'],
                        "position": emp['position'],
                        "join_date": emp.get('joinDate', ''),
                        "day_off": day_off,
                        "created_at": emp.get('createdAt', datetime.datetime.now().isoformat()),
                        "updated_at": emp.get('createdAt', datetime.datetime.now().isoformat())
                    })
                    migrated += 1
                    print(f"   Employee: {emp['name']} ({employee_no})")
                except DuplicateKeyError:
                    print(f"   Employee already exists: {emp['name']}")
                except Exception as e:
                    print(f"   Error migrating employee {emp['name']}: {e}")
            
            print(f"   Migrated {migrated} employees")
        except Exception as e:
            print(f"   Error migrating employees: {e}")
    else:
        print("\nNo employees.json found - skipping")
    
    attendance_file = os.path.join(DATA_DIR, 'attendance.json')
    if os.path.exists(attendance_file):
        print("\nMigrating attendance...")
        try:
            with open(attendance_file, 'r') as f:
                attendance_data = json.load(f)
            
            migrated = 0
            for period, period_data in attendance_data.items():
                print(f"   Processing period: {period}")
                for employee_id, employee_data in period_data.items():
                    for date, status in employee_data.items():
                        try:
                            existing = attendance_collection.find_one({
                                'period': period,
                                'employee_id': employee_id,
                                'date': date
                            })
                            
                            if existing:
                                attendance_collection.update_one(
                                    {'period': period, 'employee_id': employee_id, 'date': date},
                                    {'$set': {'status': status, 'updated_at': datetime.datetime.now().isoformat()}}
                                )
                            else:
                                attendance_collection.insert_one({
                                    "period": period,
                                    "employee_id": employee_id,
                                    "date": date,
                                    "status": status,
                                    "created_at": datetime.datetime.now().isoformat(),
                                    "updated_at": datetime.datetime.now().isoformat()
                                })
                            migrated += 1
                        except Exception as e:
                            print(f"   Error migrating attendance record {period}/{employee_id}/{date}: {e}")
                
                count = attendance_collection.count_documents({'period': period})
                print(f"   Period {period}: {count} records")
            
            print(f"   Migrated {migrated} attendance records total")
        except Exception as e:
            print(f"   Error migrating attendance: {e}")
    else:
        print("\nNo attendance.json found - skipping")
    
    print("\n" + "=" * 55)
    print("    AUTO-MIGRATION COMPLETE")
    print("=" * 55)
    try:
        print(f"\nFinal Counts:")
        print(f"   Users: {users_collection.count_documents({})}")
        print(f"   Employees: {employees_collection.count_documents({})}")
        print(f"   Attendance: {attendance_collection.count_documents({})}")
    except:
        print("   Could not retrieve counts")
    print("=" * 55 + "\n")

def generate_employee_number(department):
    prefixes = {
        'Admin & Accounts': 'EMP',
        'Front Office': 'EMP',
        'Food & Beverage': 'EMP',
        'Housekeeping & Laundry': 'EMP',
        'Maintenance': 'EMP',
        'Attachment': 'ATT',
        'Security': 'SSS'
    }
    
    prefix = prefixes.get(department, 'EMP')
    
    pattern = f"^{prefix}"
    existing = employees_collection.find({'employee_no': {'$regex': pattern}})
    
    max_num = 0
    for emp in existing:
        num_part = emp.get('employee_no', '')[len(prefix):]
        try:
            num = int(num_part)
            if num > max_num:
                max_num = num
        except:
            pass
    
    next_num = max_num + 1
    return f"{prefix}{str(next_num).zfill(4)}"

def fix_existing_employee_numbers():
    print("\nChecking for employees without numbers...")
    
    try:
        employees = employees_collection.find({
            '$or': [
                {'employee_no': {'$exists': False}},
                {'employee_no': ''},
                {'employee_no': None}
            ]
        })
        
        fixed = 0
        for emp in employees:
            department = emp.get('department', '')
            new_number = generate_employee_number(department)
            
            employees_collection.update_one(
                {'_id': emp['_id']},
                {'$set': {'employee_no': new_number}}
            )
            fixed += 1
            print(f"   Fixed: {emp.get('name')} -> {new_number}")
        
        print(f"Fixed {fixed} employees without numbers")
    except Exception as e:
        print(f"Error fixing employee numbers: {e}")

def initialize_default_users():
    try:
        if users_collection.count_documents({}) > 0:
            print("Users already exist - skipping default user creation")
            return
    except:
        print("Could not check users - skipping default user creation")
        return
    
    print("Creating default users...")
    default_users = [
        {
            "_id": "dev_001",
            "username": "developer",
            "password": "192.168.1.1",
            "name": "Timothy Kandiero",
            "role": "IT Specialist",
            "department": "IT",
            "created_at": datetime.datetime.now().isoformat()
        },
        {
            "_id": "sp_001",
            "username": "supervisor",
            "password": "super2026",
            "name": "Luckymore Kaphamtengo",
            "role": "Supervisor",
            "department": "Management",
            "created_at": datetime.datetime.now().isoformat()
        },
        {
            "_id": "admin_001",
            "username": "admin",
            "password": "smooth",
            "name": "Susan Kasola",
            "role": "System Admin",
            "department": "Admin & Accounts",
            "created_at": datetime.datetime.now().isoformat()
        },
        {
            "_id": "front_001",
            "username": "front",
            "password": "frontview",
            "name": "Front Office Viewer",
            "role": "Viewer",
            "department": "Front Office",
            "created_at": datetime.datetime.now().isoformat()
        }
    ]
    
    for user in default_users:
        try:
            users_collection.insert_one(user)
            print(f"   Default user created: {user['username']}")
        except DuplicateKeyError:
            print(f"   User already exists: {user['username']}")
        except Exception as e:
            print(f"   Error creating user {user['username']}: {e}")

def ensure_period_exists():
    try:
        periods = attendance_collection.distinct('period')
        periods = [p for p in periods if p and p != '']
        
        if not periods:
            print(" No periods found in database. Creating default period...")
            
            from datetime import datetime
            now = datetime.now()
            day = now.day
            year = now.year
            month = now.month
            if day >= 27:
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            current_period = f"{year}-{str(month).zfill(2)}"
            
            attendance_collection.insert_one({
                "period": current_period,
                "employee_id": "__init__",
                "date": current_period,
                "status": "",
                "created_at": datetime.datetime.now().isoformat(),
                "updated_at": datetime.datetime.now().isoformat()
            })
            print(f" Created default period: {current_period}")
            return current_period
        
        return periods[0]
    except Exception as e:
        print(f" Error ensuring period exists: {e}")
        return None

def check_role(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = getattr(request, 'user', None)
            if not user:
                return jsonify({"error": "Authentication required"}), 401
            
            user_role = user.get('role', '')
            if user_role not in allowed_roles:
                return jsonify({"error": "Insufficient permissions"}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def generate_token(user):
    payload = {
        "userId": user.get('_id') or user.get('username'),
        "username": user['username'],
        "role": user.get('role', 'User'),
        "department": user.get('department', ''),
        "exp": int((datetime.datetime.now() + datetime.timedelta(days=1)).timestamp() * 1000)
    }
    return base64.b64encode(json.dumps(payload).encode()).decode()

def verify_token(token):
    try:
        payload = json.loads(base64.b64decode(token.encode()).decode())
        if payload.get('exp', 0) < int(datetime.datetime.now().timestamp() * 1000):
            return None
        return payload
    except:
        return None

def auth_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({"error": "No token provided"}), 401
        
        parts = auth_header.split()
        if len(parts) != 2 or parts[0] != 'Bearer':
            return jsonify({"error": "Invalid token format"}), 401
        
        user = verify_token(parts[1])
        if not user:
            return jsonify({"error": "Invalid or expired token"}), 401
        
        token = parts[1]
        if not validate_session(token, user.get('userId')):
            return jsonify({"error": "Session expired or logged in on another device"}), 401
        
        request.user = user
        return f(*args, **kwargs)
    return decorated_function

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    
    try:
        user = users_collection.find_one({'username': username})
        
        if not user or user.get('password') != password:
            log_activity(
                user=None,
                action='LOGIN_FAILED',
                details={'username': username, 'reason': 'Invalid credentials'},
                success=False,
                ip_address=request.remote_addr
            )
            return jsonify({"error": "Invalid credentials"}), 401
        
        user.pop('password', None)
        token = generate_token(user)
        
        create_session(user.get('_id'), token)
        
        log_activity(
            user=user,
            action='LOGIN',
            details={'username': username},
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "success": True,
            "token": token,
            "user": {
                "username": user.get('username'),
                "name": user.get('name'),
                "role": user.get('role'),
                "department": user.get('department')
            }
        })
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({"error": "Login failed"}), 500

@app.route('/api/logout', methods=['POST'])
@auth_required
def logout():
    try:
        user_id = request.user.get('userId')
        token = request.headers.get('Authorization').split()[1]
        
        sessions_collection.delete_one({'token': token})
        
        log_activity(
            user=request.user,
            action='LOGOUT',
            details={'username': request.user.get('username')},
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/verify', methods=['GET'])
@auth_required
def verify():
    return jsonify({"valid": True, "user": request.user})

@app.route('/api/profile', methods=['PUT'])
@auth_required
def update_profile():
    data = request.get_json()
    user_id = request.user.get('userId')
    
    if not user_id:
        return jsonify({"error": "User not found"}), 404
    
    try:
        user = users_collection.find_one({'_id': user_id})
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        update_data = {}
        
        new_username = data.get('username')
        if new_username and new_username != user.get('username'):
            existing = users_collection.find_one({'username': new_username})
            if existing and existing['_id'] != user_id:
                return jsonify({"error": "Username already taken"}), 400
            update_data['username'] = new_username
        
        if data.get('name'):
            update_data['name'] = data.get('name')
        
        new_password = data.get('newPassword')
        if new_password:
            current_password = data.get('currentPassword')
            if not current_password:
                return jsonify({"error": "Current password required to change password"}), 400
            
            if user.get('password') != current_password:
                return jsonify({"error": "Current password is incorrect"}), 401
            
            update_data['password'] = new_password
        
        if not update_data:
            return jsonify({"error": "No fields to update"}), 400
        
        users_collection.update_one(
            {'_id': user_id},
            {'$set': update_data}
        )
        
        log_activity(
            user=request.user,
            action='PROFILE_UPDATE',
            details={'updates': list(update_data.keys())},
            ip_address=request.remote_addr
        )
        
        updated_user = users_collection.find_one({'_id': user_id})
        updated_user.pop('password', None)
        updated_user['_id'] = str(updated_user['_id'])
        
        return jsonify({
            "success": True,
            "message": "Profile updated successfully",
            "user": {
                "username": updated_user.get('username'),
                "name": updated_user.get('name'),
                "role": updated_user.get('role'),
                "department": updated_user.get('department')
            }
        })
        
    except Exception as e:
        print(f"Error updating profile: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees', methods=['GET'])
@auth_required
def get_employees():
    try:
        dept_order = ['Admin & Accounts', 'Front Office', 'Food & Beverage', 
                      'Housekeeping & Laundry', 'Maintenance', 'Attachment', 'Security']
        
        employees = list(employees_collection.find())
        employees.sort(key=lambda x: (dept_order.index(x.get('department')) if x.get('department') in dept_order else 999, x.get('name', '')))
        
        for emp in employees:
            emp['_id'] = str(emp['_id'])
            if 'day_off' not in emp:
                emp['day_off'] = ''
        
        return jsonify(employees)
    except Exception as e:
        print(f"Error fetching employees: {e}")
        return jsonify({"error": "Failed to fetch employees"}), 500

@app.route('/api/employees', methods=['POST'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def add_employee():
    data = request.get_json()
    
    name = data.get('name')
    department = data.get('department')
    position = data.get('position')
    day_off = data.get('day_off', '')
    
    if not name or not department or not position:
        return jsonify({"error": "Name, department, and position are required"}), 400
    
    try:
        employee_no = generate_employee_number(department)
        
        new_employee = {
            "_id": str(int(datetime.datetime.now().timestamp() * 1000)),
            "employee_no": employee_no,
            "name": name.upper(),
            "department": department,
            "position": position.upper(),
            "join_date": data.get('joinDate', ''),
            "day_off": day_off,
            "created_at": datetime.datetime.now().isoformat(),
            "updated_at": datetime.datetime.now().isoformat()
        }
        
        employees_collection.insert_one(new_employee)
        new_employee['_id'] = str(new_employee['_id'])
        
        log_activity(
            user=request.user,
            action='EMPLOYEE_ADD',
            details={
                'employee_id': new_employee['_id'],
                'name': new_employee['name'],
                'department': new_employee['department'],
                'position': new_employee['position'],
                'employee_no': new_employee['employee_no'],
                'day_off': day_off
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True, "employee": new_employee})
    except DuplicateKeyError:
        return jsonify({"error": "Employee number already exists"}), 400
    except Exception as e:
        print(f"Error adding employee: {e}")
        return jsonify({"error": f"Failed to add employee: {str(e)}"}), 500

@app.route('/api/employees/<employee_id>', methods=['PUT'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def update_employee(employee_id):
    data = request.get_json()
    
    update_data = {
        "name": data.get('name', '').upper(),
        "department": data.get('department', ''),
        "position": data.get('position', '').upper(),
        "employee_no": data.get('employeeNo', ''),
        "join_date": data.get('joinDate', ''),
        "day_off": data.get('day_off', ''),
        "updated_at": datetime.datetime.now().isoformat()
    }
    
    update_data = {k: v for k, v in update_data.items() if v}
    
    try:
        result = employees_collection.update_one(
            {'_id': employee_id},
            {'$set': update_data}
        )
        
        if result.matched_count == 0:
            return jsonify({"error": "Employee not found"}), 404
        
        updated_employee = employees_collection.find_one({'_id': employee_id})
        updated_employee['_id'] = str(updated_employee['_id'])
        
        log_activity(
            user=request.user,
            action='EMPLOYEE_UPDATE',
            details={
                'employee_id': employee_id,
                'updates': update_data
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True, "employee": updated_employee})
    except Exception as e:
        print(f"Error updating employee: {e}")
        return jsonify({"error": "Failed to update employee"}), 500

@app.route('/api/employees/<employee_id>', methods=['DELETE'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def delete_employee(employee_id):
    try:
        print(f"Deleting employee: {employee_id}")
        
        result = employees_collection.delete_one({'_id': employee_id})
        
        if result.deleted_count == 0:
            print(f"   Employee not found: {employee_id}")
            return jsonify({"error": "Employee not found"}), 404
        
        att_result = attendance_collection.delete_many({'employee_id': employee_id})
        print(f"   Deleted {att_result.deleted_count} attendance records")
        
        log_activity(
            user=request.user,
            action='EMPLOYEE_DELETE',
            details={
                'employee_id': employee_id
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True, "message": "Employee deleted successfully"})
    except Exception as e:
        print(f"Error deleting employee: {e}")
        return jsonify({"error": f"Failed to delete employee: {str(e)}"}), 500

@app.route('/api/attendance/<period>', methods=['GET'])
@auth_required
def get_attendance(period):
    try:
        print(f" Fetching attendance for period: {period}")
        
        auto_mark_day_offs(period)
        
        records = attendance_collection.find({'period': period})
        
        formatted_data = {}
        for rec in records:
            emp_id = rec.get('employee_id')
            date = rec.get('date')
            status = rec.get('status', '')
            
            if not emp_id or not date or emp_id == '__init__':
                continue
                
            if emp_id not in formatted_data:
                formatted_data[emp_id] = {}
            formatted_data[emp_id][date] = status
        
        print(f" Returning data for {len(formatted_data)} employees")
        return jsonify(formatted_data)
        
    except Exception as e:
        print(f" Error fetching attendance: {e}")
        return jsonify({"error": "Failed to fetch attendance"}), 500

@app.route('/api/attendance', methods=['POST'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def save_attendance():
    data = request.get_json()
    
    period = data.get('period')
    employee_id = data.get('employeeId')
    date = data.get('date')
    status = data.get('status', '')
    
    if not period or not employee_id or not date:
        return jsonify({"error": "Missing required fields"}), 400
    
    try:
        print(f" Saving attendance: period={period}, employee={employee_id}, date={date}, status={status}")
        
        attendance_collection.update_one(
            {'period': period, 'employee_id': employee_id, 'date': date},
            {'$set': {
                'status': status,
                'updated_at': datetime.datetime.now().isoformat()
            }},
            upsert=True
        )
        
        log_activity(
            user=request.user,
            action='ATTENDANCE_UPDATE',
            details={
                'period': period,
                'employee_id': employee_id,
                'date': date,
                'status': status
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True})
    except Exception as e:
        print(f" Error saving attendance: {e}")
        return jsonify({"error": f"Failed to save attendance: {str(e)}"}), 500

@app.route('/api/attendance/mark-dayoffs/<period>', methods=['POST'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def mark_day_offs(period):
    try:
        count = auto_mark_day_offs(period)
        
        return jsonify({
            "success": True,
            "period": period,
            "marked_count": count,
            "message": f"Marked {count} day-off records"
        })
    except Exception as e:
        print(f" Error marking day-offs: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/periods', methods=['GET'])
@auth_required
def get_periods():
    try:
        periods = attendance_collection.distinct('period')
        periods = [p for p in periods if p and p != '' and p != '__init__']
        periods.sort(reverse=True)
        
        print(f" Found periods: {periods}")
        
        if not periods:
            from datetime import datetime
            now = datetime.now()
            day = now.day
            year = now.year
            month = now.month
            if day >= 27:
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            current_period = f"{year}-{str(month).zfill(2)}"
            
            attendance_collection.insert_one({
                "period": current_period,
                "employee_id": "__init__",
                "date": current_period,
                "status": "",
                "created_at": datetime.datetime.now().isoformat(),
                "updated_at": datetime.datetime.now().isoformat()
            })
            periods = [current_period]
            print(f" Created current period: {current_period}")
        
        return jsonify(periods)
    except Exception as e:
        print(f" Error fetching periods: {e}")
        from datetime import datetime
        now = datetime.now()
        day = now.day
        year = now.year
        month = now.month
        if day >= 27:
            month += 1
            if month > 12:
                month = 1
                year += 1
        current_period = f"{year}-{str(month).zfill(2)}"
        return jsonify([current_period])

@app.route('/api/employees/light', methods=['GET'])
@auth_required
def get_employees_light():
    try:
        employees = list(employees_collection.find({}, {
            '_id': 1,
            'name': 1,
            'department': 1,
            'position': 1,
            'employee_no': 1
        }))
        
        dept_order = ['Admin & Accounts', 'Front Office', 'Food & Beverage', 
                      'Housekeeping & Laundry', 'Maintenance', 'Attachment', 'Security']
        
        employees.sort(key=lambda x: (dept_order.index(x.get('department')) if x.get('department') in dept_order else 999, x.get('name', '')))
        
        for emp in employees:
            emp['_id'] = str(emp['_id'])
        
        return jsonify(employees)
    except Exception as e:
        print(f"Error fetching employees: {e}")
        return jsonify({"error": "Failed to fetch employees"}), 500

@app.route('/api/employee-summary/<period>', methods=['GET'])
@auth_required
def get_employee_summary(period):
    try:
        employees = list(employees_collection.find({}, {
            '_id': 1,
            'name': 1,
            'department': 1,
            'position': 1,
            'employee_no': 1,
            'join_date': 1
        }))
        
        attendance_records = attendance_collection.find({'period': period})
        
        attendance_data = {}
        for rec in attendance_records:
            emp_id = rec.get('employee_id')
            date = rec.get('date')
            status = rec.get('status', '')
            if emp_id and emp_id != '__init__' and date:
                if emp_id not in attendance_data:
                    attendance_data[emp_id] = {}
                attendance_data[emp_id][date] = status
        
        summary = []
        for emp in employees:
            emp_id = emp['_id']
            emp_att = attendance_data.get(emp_id, {})
            
            p = 0
            a = 0
            s = 0
            l = 0
            o = 0
            
            for status in emp_att.values():
                if status == 'P':
                    p += 1
                elif status == 'A':
                    a += 1
                elif status == 'S':
                    s += 1
                elif status == 'L':
                    l += 1
                elif status == 'O':
                    o += 1
            
            total = p + a + s + l + o
            if total > 0:
                summary.append({
                    'employee_id': str(emp['_id']),
                    'employee_no': emp.get('employee_no', ''),
                    'name': emp.get('name', ''),
                    'department': emp.get('department', ''),
                    'position': emp.get('position', ''),
                    'join_date': emp.get('join_date', ''),
                    'present': p,
                    'absent': a,
                    'sick': s,
                    'leave': l,
                    'off': o,
                    'total_days': total
                })
        
        return jsonify({
            'success': True,
            'period': period,
            'summary': summary
        })
    except Exception as e:
        print(f"Error getting employee summary: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/employee-summary-range', methods=['POST'])
@auth_required
def get_employee_summary_range():
    try:
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"error": "Start date and end date are required"}), 400
        
        employees = list(employees_collection.find({}, {
            '_id': 1,
            'name': 1,
            'department': 1,
            'position': 1,
            'employee_no': 1,
            'join_date': 1
        }))
        
        attendance_records = attendance_collection.find({
            'date': {'$gte': start_date, '$lte': end_date}
        })
        
        attendance_data = {}
        for rec in attendance_records:
            emp_id = rec.get('employee_id')
            date = rec.get('date')
            status = rec.get('status', '')
            if emp_id and emp_id != '__init__' and date:
                if emp_id not in attendance_data:
                    attendance_data[emp_id] = {}
                attendance_data[emp_id][date] = status
        
        summary = []
        for emp in employees:
            emp_id = emp['_id']
            emp_att = attendance_data.get(emp_id, {})
            
            p = 0
            a = 0
            s = 0
            l = 0
            o = 0
            
            for status in emp_att.values():
                if status == 'P':
                    p += 1
                elif status == 'A':
                    a += 1
                elif status == 'S':
                    s += 1
                elif status == 'L':
                    l += 1
                elif status == 'O':
                    o += 1
            
            total = p + a + s + l + o
            if total > 0:
                summary.append({
                    'employee_id': str(emp['_id']),
                    'employee_no': emp.get('employee_no', ''),
                    'name': emp.get('name', ''),
                    'department': emp.get('department', ''),
                    'position': emp.get('position', ''),
                    'join_date': emp.get('join_date', ''),
                    'present': p,
                    'absent': a,
                    'sick': s,
                    'leave': l,
                    'off': o,
                    'total_days': total
                })
        
        return jsonify({
            'success': True,
            'start_date': start_date,
            'end_date': end_date,
            'summary': summary
        })
    except Exception as e:
        print(f"Error getting employee summary range: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/report/<period>', methods=['GET'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def export_report(period):
    try:
        departments = ['Admin & Accounts', 'Front Office', 'Food & Beverage', 
                      'Housekeeping & Laundry', 'Maintenance', 'Attachment', 'Security']
        
        employees = list(employees_collection.find())
        employees.sort(key=lambda x: (departments.index(x.get('department')) if x.get('department') in departments else 999, x.get('name', '')))
        
        attendance_records = attendance_collection.find({'period': period})
        attendance_data = {}
        for record in attendance_records:
            emp_id = record.get('employee_id')
            date = record.get('date')
            status = record.get('status', '')
            if emp_id and emp_id != '__init__':
                if emp_id not in attendance_data:
                    attendance_data[emp_id] = {}
                attendance_data[emp_id][date] = status
        
        year, month = map(int, period.split('-'))
        prev_month = 12 if month == 1 else month - 1
        prev_year = year - 1 if month == 1 else year
        
        dates = []
        import calendar
        days_in_prev = calendar.monthrange(prev_year, prev_month)[1]
        for d in range(27, days_in_prev + 1):
            date_str = f"{prev_year}-{str(prev_month).zfill(2)}-{str(d).zfill(2)}"
            dates.append({
                'date': date_str,
                'label': f"{prev_month}/{d}"
            })
        
        days_in_curr = calendar.monthrange(year, month)[1]
        for d in range(1, min(27, days_in_curr + 1)):
            date_str = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"
            dates.append({
                'date': date_str,
                'label': f"{month}/{d}"
            })
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'
        
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        header_font = Font(bold=True, size=9, color="FFFFFF")
        header_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        dept_font = Font(bold=True, size=10, color="FFFFFF")
        dept_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        dept_alignment = Alignment(horizontal="left", vertical="center")
        
        total_font = Font(bold=True, size=9)
        total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        status_colors = {
            'P': '92D050',
            'A': 'FF0000',
            'S': 'FFC000',
            'L': '00B0F0',
            'O': 'D9D9D9'
        }
        
        total_cols = len(dates) + 3 + 5
        last_col = get_column_letter(total_cols)
        
        ws.merge_cells(f'A1:{last_col}1')
        title_cell = ws['A1']
        title_cell.value = 'SILVER SANDS SALIMA - HUMAN RESOURCE MANAGEMENT SYSTEM'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells(f'A2:{last_col}2')
        sub_cell = ws['A2']
        month_name = datetime.datetime(year, month, 1).strftime('%B')
        prev_month_name = datetime.datetime(prev_year, prev_month, 1).strftime('%B')
        sub_cell.value = f"ATTENDANCE REPORT — Payroll Period: 27 {prev_month_name} {prev_year} to 26 {month_name} {year}"
        sub_cell.font = Font(bold=True, size=11, color="FFFFFF")
        sub_cell.fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 5
        
        header_row = 5
        headers = ['#', 'Employee Name', 'Position']
        headers.extend([d['label'] for d in dates])
        headers.extend(['P', 'A', 'S', 'L', 'O'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[header_row].height = 40
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        for i in range(4, len(dates) + 4):
            ws.column_dimensions[get_column_letter(i)].width = 4.5
        for i in range(len(dates) + 4, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 7
        
        row_num = 6
        grand_p = grand_a = grand_s = grand_l = grand_o = 0        
        for dept in departments:
            dept_emps = [e for e in employees if e.get('department') == dept]
            if not dept_emps:
                continue
            
            ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
            dept_cell = ws.cell(row=row_num, column=1)
            dept_cell.value = f"  {dept.upper()}"
            dept_cell.font = dept_font
            dept_cell.fill = dept_fill
            dept_cell.alignment = dept_alignment
            dept_cell.border = thin_border
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            
            dept_p = dept_a = dept_s = dept_l = dept_o = 0
            
            for emp_idx, emp in enumerate(dept_emps, 1):
                emp_attendance = attendance_data.get(emp['_id'], {})
                
                ws.cell(row=row_num, column=1, value=emp_idx)
                ws.cell(row=row_num, column=2, value=emp.get('name', ''))
                ws.cell(row=row_num, column=3, value=emp.get('position', ''))
                
                for col in range(1, 4):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = Font(size=9)
                    cell.border = thin_border
                    if col == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                p = a = s = l = o = 0
                
                for date_idx, date_info in enumerate(dates, 4):
                    status = emp_attendance.get(date_info['date'], '')
                    cell = ws.cell(row=row_num, column=date_idx)
                    cell.value = status
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=8, bold=True)
                    
                    if status and status in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[status],
                            end_color=status_colors[status],
                            fill_type="solid"
                        )
                    
                    if status == 'P':
                        p += 1
                    elif status == 'A':
                        a += 1
                    elif status == 'S':
                        s += 1
                    elif status == 'L':
                        l += 1
                    elif status == 'O':
                        o += 1
                
                dept_p += p
                dept_a += a
                dept_s += s
                dept_l += l
                dept_o += o
                
                totals = [p, a, s, l, o]
                for total_idx, total_val in enumerate(totals, len(dates) + 4):
                    cell = ws.cell(row=row_num, column=total_idx)
                    cell.value = total_val
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=9, bold=True)
                
                ws.row_dimensions[row_num].height = 16
                row_num += 1
            
            ws.cell(row=row_num, column=2, value=f"{dept} Totals")
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
            
            for i in range(4, len(dates) + 4):
                cell = ws.cell(row=row_num, column=i)
                cell.fill = total_fill
                cell.border = thin_border
            
            dept_totals = [dept_p, dept_a, dept_s, dept_l, dept_o]
            for total_idx, total_val in enumerate(dept_totals, len(dates) + 4):
                cell = ws.cell(row=row_num, column=total_idx)
                cell.value = total_val
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
            
            ws.row_dimensions[row_num].height = 16
            row_num += 1
            
            grand_p += dept_p
            grand_a += dept_a
            grand_s += dept_s
            grand_l += dept_l
            grand_o += dept_o
        
        ws.merge_cells(f'A{row_num}:{get_column_letter(len(dates) + 3)}{row_num}')
        grand_cell = ws.cell(row=row_num, column=1)
        grand_cell.value = 'GRAND TOTALS'
        grand_cell.font = Font(bold=True, size=10, color="FFFFFF")
        grand_cell.fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        grand_cell.alignment = Alignment(horizontal="center", vertical="center")
        grand_cell.border = thin_border
        
        grand_totals = [grand_p, grand_a, grand_s, grand_l, grand_o]
        for total_idx, total_val in enumerate(grand_totals, len(dates) + 4):
            cell = ws.cell(row=row_num, column=total_idx)
            cell.value = total_val
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
            cell.border = thin_border
        
        ws.row_dimensions[row_num].height = 20
        
        ws.protection.password = 'hrprotect2024'
        ws.protection.sheet = True
        
        log_activity(
            user=request.user,
            action='REPORT_DOWNLOAD',
            details={'period': period},
            ip_address=request.remote_addr
        )
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Report_{period}.xlsx'
        )
        
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500

@app.route('/api/report/stats/<period>', methods=['GET'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def get_report_stats(period):
    try:
        attendance_records = attendance_collection.find({'period': period})
        total_p = 0
        total_a = 0
        total_s = 0
        total_l = 0
        total_o = 0
        
        for record in attendance_records:
            status = record.get('status', '')
            if status == 'P':
                total_p += 1
            elif status == 'A':
                total_a += 1
            elif status == 'S':
                total_s += 1
            elif status == 'L':
                total_l += 1
            elif status == 'O':
                total_o += 1
        
        return jsonify({
            "success": True,
            "period": period,
            "stats": {
                "present": total_p,
                "absent": total_a,
                "sick": total_s,
                "leave": total_l,
                "off": total_o,
                "total": total_p + total_a + total_s + total_l + total_o
            }
        })
    except Exception as e:
        print(f"Error getting report stats: {e}")
        return jsonify({"error": "Failed to get stats"}), 500

@app.route('/api/activity-logs', methods=['GET'])
@auth_required
def get_activity_logs_route():
    try:
        limit = int(request.args.get('limit', 100))
        skip = int(request.args.get('skip', 0))
        action_filter = request.args.get('action', None)
        username_filter = request.args.get('username', None)
        date_from = request.args.get('date_from', None)
        date_to = request.args.get('date_to', None)
        
        filters = {}
        if action_filter:
            filters['action'] = action_filter
        if username_filter:
            filters['username'] = username_filter
        if date_from or date_to:
            filters['timestamp'] = {}
            if date_from:
                filters['timestamp']['$gte'] = date_from
            if date_to:
                filters['timestamp']['$lte'] = date_to + 'T23:59:59.999Z'
        
        user_obj = request.user
        logs = get_activity_logs(user_obj, filters, limit, skip)
        total = get_activity_logs_count(user_obj, filters)
        
        if request.user.get('role') == 'IT Specialist':
            distinct_actions = activity_log_collection.distinct('action')
            distinct_users = activity_log_collection.distinct('username')
        else:
            distinct_actions = activity_log_collection.distinct('action', {'user_id': request.user.get('userId')})
            distinct_users = [request.user.get('username')]
        
        return jsonify({
            "success": True,
            "logs": logs,
            "total": total,
            "limit": limit,
            "skip": skip,
            "filters": {
                "actions": distinct_actions,
                "users": distinct_users
            }
        })
    except Exception as e:
        print(f"Error fetching activity logs: {e}")
        return jsonify({"error": "Failed to fetch activity logs"}), 500

@app.route('/api/activity-logs/<log_id>', methods=['DELETE'])
@auth_required
@check_role(['IT Specialist'])
def delete_activity_log_route(log_id):
    try:
        success = delete_activity_log(log_id)
        if success:
            log_activity(
                user=request.user,
                action='LOG_DELETE',
                details={'log_id': log_id},
                ip_address=request.remote_addr
            )
            return jsonify({"success": True, "message": "Log deleted successfully"})
        else:
            return jsonify({"error": "Log not found"}), 404
    except Exception as e:
        print(f"Error deleting log: {e}")
        return jsonify({"error": "Failed to delete log"}), 500

@app.route('/api/activity-logs/clear', methods=['POST'])
@auth_required
@check_role(['IT Specialist'])
def clear_activity_logs_route():
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        
        if not filters or (filters.get('action') == 'SYSTEM' and not data.get('confirm')):
            return jsonify({
                "error": "Confirmation required to clear all logs",
                "total": activity_log_collection.count_documents({})
            }), 400
        
        deleted_count = clear_activity_logs(filters)
        
        log_activity(
            user=request.user,
            action='LOGS_CLEARED',
            details={'deleted_count': deleted_count, 'filters': filters},
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "success": True,
            "message": f"Cleared {deleted_count} logs",
            "deleted_count": deleted_count
        })
    except Exception as e:
        print(f"Error clearing logs: {e}")
        return jsonify({"error": "Failed to clear logs"}), 500

@app.route('/api/users', methods=['GET'])
@auth_required
@check_role(['IT Specialist', 'System Admin', 'Supervisor'])
def get_users():
    try:
        users = list(users_collection.find({}, {
            '_id': 1,
            'username': 1,
            'name': 1,
            'role': 1,
            'department': 1,
            'created_at': 1
        }))
        
        for user in users:
            user['_id'] = str(user['_id'])
        
        return jsonify(users)
    except Exception as e:
        print(f"Error fetching users: {e}")
        return jsonify({"error": "Failed to fetch users"}), 500

@app.route('/api/users', methods=['POST'])
@auth_required
@check_role(['IT Specialist', 'Supervisor'])
def create_user():
    data = request.get_json()
    
    username = data.get('username')
    name = data.get('name')
    password = data.get('password')
    role = data.get('role', 'Viewer')
    department = data.get('department', '')
    
    if not username:
        return jsonify({"error": "Username is required"}), 400
    if len(username) < 3:
        return jsonify({"error": "Username must be at least 3 characters"}), 400
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not password:
        return jsonify({"error": "Password is required"}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400
    
    existing = users_collection.find_one({'username': username})
    if existing:
        return jsonify({"error": "Username already exists"}), 400
    
    try:
        new_user = {
            "_id": str(int(datetime.datetime.now().timestamp() * 1000)),
            "username": username,
            "password": password,
            "name": name,
            "role": role,
            "department": department,
            "created_at": datetime.datetime.now().isoformat()
        }
        
        users_collection.insert_one(new_user)
        
        log_activity(
            user=request.user,
            action='USER_CREATE',
            details={
                'username': username,
                'name': name,
                'role': role,
                'department': department
            },
            ip_address=request.remote_addr
        )
        
        new_user.pop('password', None)
        new_user['_id'] = str(new_user['_id'])
        
        return jsonify({
            "success": True,
            "message": "User created successfully",
            "user": new_user
        })
        
    except DuplicateKeyError:
        return jsonify({"error": "Username already exists"}), 400
    except Exception as e:
        print(f"Error creating user: {e}")
        return jsonify({"error": f"Failed to create user: {str(e)}"}), 500

@app.route('/api/users/<user_id>', methods=['PUT'])
@auth_required
@check_role(['IT Specialist'])
def update_user(user_id):
    data = request.get_json()
    
    update_data = {}
    if 'role' in data:
        update_data['role'] = data['role']
    if 'department' in data:
        update_data['department'] = data['department']
    if 'name' in data:
        update_data['name'] = data['name']
    
    if not update_data:
        return jsonify({"error": "No fields to update"}), 400
    
    try:
        result = users_collection.update_one(
            {'_id': user_id},
            {'$set': update_data}
        )
        
        if result.matched_count == 0:
            return jsonify({"error": "User not found"}), 404
        
        log_activity(
            user=request.user,
            action='USER_UPDATE',
            details={
                'user_id': user_id,
                'updates': update_data
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True, "message": "User updated successfully"})
    except Exception as e:
        print(f"Error updating user: {e}")
        return jsonify({"error": "Failed to update user"}), 500

@app.route('/api/users/<user_id>', methods=['DELETE'])
@auth_required
@check_role(['IT Specialist'])
def delete_user(user_id):
    try:
        user = users_collection.find_one({'_id': user_id})
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        if user.get('username') == 'developer':
            return jsonify({"error": "Cannot delete the developer account"}), 403
        
        result = users_collection.delete_one({'_id': user_id})
        
        if result.deleted_count == 0:
            return jsonify({"error": "User not found"}), 404
        
        clear_user_sessions(user_id)
        
        log_activity(
            user=request.user,
            action='USER_DELETE',
            details={'user_id': user_id},
            ip_address=request.remote_addr
        )
        
        return jsonify({"success": True, "message": "User deleted successfully"})
    except Exception as e:
        print(f"Error deleting user: {e}")
        return jsonify({"error": "Failed to delete user"}), 500

# ============================================================
# AI ASSISTANT ROUTES - Works with both old and new OpenAI
# ============================================================

def call_openai(messages, temperature=0.3, max_tokens=300):
    """Universal OpenAI caller that works with both old and new versions"""
    if not OPENAI_AVAILABLE:
        return None
    
    try:
        # Try new version first
        if hasattr(openai_client, 'chat') and hasattr(openai_client.chat, 'completions'):
            response = openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=messages,
                temperature=temperature,
                max_tokens=max_tokens
            )
            return response.choices[0].message.content.strip()
        # Try old version
        elif hasattr(openai_client, 'ChatCompletion'):
            response = openai_client.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages,
                temperature=temperature,
                max_tokens=max_tokens
            )
            return response.choices[0].message.content.strip()
        else:
            return None
    except Exception as e:
        print(f"OpenAI call failed: {e}")
        return None

@app.route('/api/ai/ask', methods=['POST'])
@auth_required
def ai_ask():
    try:
        data = request.get_json()
        question = data.get('question', '').strip()
        
        if not question:
            return jsonify({
                "success": False,
                "message": "Please ask a question",
                "type": "error"
            })
        
        # Try OpenAI first if available
        if OPENAI_AVAILABLE:
            try:
                current_date = datetime.datetime.now().strftime("%Y-%m-%d")
                current_period = get_current_period()
                
                system_prompt = f"""You are an HR assistant for Silver Sands Salima.

The database contains:
- Employees: name, department, position, employee_no, join_date, day_off
- Attendance: period, employee_id, date, status (P=Present, A=Absent, S=Sick, L=Leave, O=Day Off)

Current date: {current_date}
Current period: {current_period}

Available operations:
1. get_employees_by_department(department)
2. get_employee_by_name(name)
3. get_attendance_summary(period, department)
4. get_attendance_stats(period)
5. get_department_attendance(period)
6. get_absent_employees(period, min_days)
7. get_best_attendance(period, limit)
8. get_worst_attendance(period, limit)
9. get_employees_by_day_off(day)
10. get_all_employees()
11. get_employee_attendance_totals(employee_name)

Respond ONLY with a JSON object in this exact format:
{{"operation": "operation_name", "params": {{"param1": "value1"}}, "explanation": "brief explanation"}}

If the question is NOT about HR data:
{{"operation": "general", "message": "I'm here to help with HR data."}}"""

                ai_text = call_openai([
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": question}
                ])
                
                if ai_text:
                    try:
                        json_match = re.search(r'\{.*\}', ai_text, re.DOTALL)
                        if json_match:
                            action = json.loads(json_match.group())
                        else:
                            action = json.loads(ai_text)
                        
                        if action.get('operation') == 'general':
                            return jsonify({
                                "success": True,
                                "message": action.get('message', "I'm here to help with HR data."),
                                "type": "general"
                            })
                        
                        result = execute_ai_operation(action)
                        return jsonify(result)
                    except Exception as e:
                        print(f"JSON parsing error: {e}")
                        # Fall through to keyword matching
            except Exception as e:
                print(f"OpenAI failed, using fallback: {e}")
                # Fall through to keyword matching
        
        # ============================================================
        # KEYWORD MATCHING FALLBACK - WORKS WITHOUT ANY API KEY
        # ============================================================
        question_lower = question.lower()
        
        if 'attendance' in question_lower and 'summary' in question_lower:
            period = get_current_period()
            result = execute_ai_operation({
                "operation": "get_attendance_summary",
                "params": {"period": period}
            })
            return jsonify(result)
        
        elif 'absent' in question_lower:
            period = get_current_period()
            result = execute_ai_operation({
                "operation": "get_absent_employees",
                "params": {"period": period, "min_days": 1}
            })
            return jsonify(result)
        
        elif 'employee' in question_lower and ('list' in question_lower or 'all' in question_lower):
            result = execute_ai_operation({
                "operation": "get_all_employees",
                "params": {}
            })
            return jsonify(result)
        
        elif 'department' in question_lower:
            departments = ['admin', 'front office', 'food', 'housekeeping', 'maintenance', 'attachment', 'security']
            dept_found = None
            for dept in departments:
                if dept in question_lower:
                    dept_found = dept.title()
                    break
            
            if dept_found:
                result = execute_ai_operation({
                    "operation": "get_employees_by_department",
                    "params": {"department": dept_found}
                })
                return jsonify(result)
            else:
                result = execute_ai_operation({
                    "operation": "get_department_attendance",
                    "params": {"period": get_current_period()}
                })
                return jsonify(result)
        
        elif 'best' in question_lower or 'top' in question_lower:
            period = get_current_period()
            result = execute_ai_operation({
                "operation": "get_best_attendance",
                "params": {"period": period, "limit": 5}
            })
            return jsonify(result)
        
        elif 'worst' in question_lower or 'bottom' in question_lower:
            period = get_current_period()
            result = execute_ai_operation({
                "operation": "get_worst_attendance",
                "params": {"period": period, "limit": 5}
            })
            return jsonify(result)
        
        elif 'day off' in question_lower or 'off' in question_lower:
            days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            day_found = None
            for day in days:
                if day in question_lower:
                    day_found = day.title()
                    break
            
            if day_found:
                result = execute_ai_operation({
                    "operation": "get_employees_by_day_off",
                    "params": {"day": day_found}
                })
                return jsonify(result)
            else:
                return jsonify({
                    "success": False,
                    "message": "Please specify a day (Monday, Tuesday, etc.)",
                    "type": "error"
                })
        
        elif 'employee' in question_lower and 'attendance' in question_lower:
            words = question.split()
            common_words = ['show', 'me', 'for', 'the', 'of', 'in', 'at', 'employee', 'attendance', 'totals', 'summary', 'view', 'get', 'see']
            name_candidates = [w for w in words if w.lower() not in common_words and len(w) > 2]
            if name_candidates:
                name = ' '.join(name_candidates)
                result = execute_ai_operation({
                    "operation": "get_employee_attendance_totals",
                    "params": {"employee_name": name}
                })
                return jsonify(result)
            else:
                return jsonify({
                    "success": False,
                    "message": "Please specify an employee name.",
                    "type": "error"
                })
        
        else:
            return jsonify({
                "success": False,
                "message": "Try: 'Show me attendance summary', 'Who was absent today?', 'List Front Office employees'",
                "type": "error"
            })
        
    except Exception as e:
        print(f"AI API Error: {e}")
        return jsonify({
            "success": False,
            "message": "Something went wrong. Please try again later.",
            "type": "error"
        })

def execute_ai_operation(action):
    operation = action.get('operation')
    params = action.get('params', {})
    
    try:
        if operation == 'get_employees_by_department':
            department = params.get('department', '').strip()
            if not department:
                return {
                    "success": False,
                    "message": "Please specify a department name.",
                    "type": "error"
                }
            
            employees = list(employees_collection.find({
                'department': {'$regex': department, '$options': 'i'}
            }))
            
            if employees:
                return {
                    "success": True,
                    "message": f"Found {len(employees)} employees in {department.title()}:",
                    "data": employees,
                    "type": "list"
                }
            else:
                return {
                    "success": True,
                    "message": f"No employees found in {department.title()}",
                    "data": [],
                    "type": "empty"
                }
        
        elif operation == 'get_employee_by_name':
            name = params.get('name', '').strip()
            if not name:
                return {
                    "success": False,
                    "message": "Please specify an employee name.",
                    "type": "error"
                }
            
            employees = list(employees_collection.find({
                'name': {'$regex': name, '$options': 'i'}
            }))
            
            if employees:
                return {
                    "success": True,
                    "message": f"Found {len(employees)} employees matching '{name}':",
                    "data": employees,
                    "type": "list"
                }
            else:
                return {
                    "success": True,
                    "message": f"No employees found matching '{name}'",
                    "data": [],
                    "type": "empty"
                }
        
        elif operation == 'get_attendance_summary':
            period = params.get('period', get_current_period())
            department = params.get('department')
            
            query = {}
            if department:
                query['department'] = {'$regex': department, '$options': 'i'}
            employees = list(employees_collection.find(query))
            
            att_records = list(attendance_collection.find({'period': period}))
            att_data = {}
            for rec in att_records:
                emp_id = rec.get('employee_id')
                if emp_id and emp_id != '__init__':
                    if emp_id not in att_data:
                        att_data[emp_id] = {}
                    att_data[emp_id][rec.get('date')] = rec.get('status', '')
            
            total_p = total_a = total_s = total_l = total_o = 0
            total_employees = len(employees)
            
            for emp in employees:
                emp_att = att_data.get(emp['_id'], {})
                for status in emp_att.values():
                    if status == 'P': total_p += 1
                    elif status == 'A': total_a += 1
                    elif status == 'S': total_s += 1
                    elif status == 'L': total_l += 1
                    elif status == 'O': total_o += 1
            
            total_days = total_p + total_a + total_s + total_l + total_o
            attendance_rate = round((total_p / total_days * 100) if total_days > 0 else 0, 1)
            
            dept_text = f" in {department.title()}" if department else ""
            
            return {
                "success": True,
                "message": f"Attendance summary for {period}{dept_text}:",
                "data": {
                    "period": period,
                    "department": department,
                    "total_employees": total_employees,
                    "present": total_p,
                    "absent": total_a,
                    "sick": total_s,
                    "leave": total_l,
                    "day_off": total_o,
                    "total_days": total_days,
                    "attendance_rate": attendance_rate
                },
                "type": "stats"
            }
        
        elif operation == 'get_attendance_stats':
            period = params.get('period', get_current_period())
            
            att_records = list(attendance_collection.find({'period': period}))
            total_p = total_a = total_s = total_l = total_o = 0
            
            for rec in att_records:
                status = rec.get('status', '')
                if status == 'P': total_p += 1
                elif status == 'A': total_a += 1
                elif status == 'S': total_s += 1
                elif status == 'L': total_l += 1
                elif status == 'O': total_o += 1
            
            total_days = total_p + total_a + total_s + total_l + total_o
            attendance_rate = round((total_p / total_days * 100) if total_days > 0 else 0, 1)
            
            return {
                "success": True,
                "message": f"Attendance statistics for {period}:",
                "data": {
                    "period": period,
                    "present": total_p,
                    "absent": total_a,
                    "sick": total_s,
                    "leave": total_l,
                    "day_off": total_o,
                    "total_days": total_days,
                    "attendance_rate": attendance_rate
                },
                "type": "stats"
            }
        
        elif operation == 'get_department_attendance':
            period = params.get('period', get_current_period())
            
            employees = list(employees_collection.find())
            att_records = list(attendance_collection.find({'period': period}))
            
            att_data = {}
            for rec in att_records:
                emp_id = rec.get('employee_id')
                if emp_id and emp_id != '__init__':
                    if emp_id not in att_data:
                        att_data[emp_id] = {}
                    att_data[emp_id][rec.get('date')] = rec.get('status', '')
            
            dept_stats = {}
            for emp in employees:
                dept = emp.get('department', 'Unknown')
                if dept not in dept_stats:
                    dept_stats[dept] = {'p': 0, 'a': 0, 's': 0, 'l': 0, 'o': 0}
                
                emp_att = att_data.get(emp['_id'], {})
                for status in emp_att.values():
                    if status == 'P': dept_stats[dept]['p'] += 1
                    elif status == 'A': dept_stats[dept]['a'] += 1
                    elif status == 'S': dept_stats[dept]['s'] += 1
                    elif status == 'L': dept_stats[dept]['l'] += 1
                    elif status == 'O': dept_stats[dept]['o'] += 1
            
            return {
                "success": True,
                "message": f"Department attendance for {period}:",
                "data": dept_stats,
                "type": "departments"
            }
        
        elif operation == 'get_absent_employees':
            period = params.get('period', get_current_period())
            min_days = params.get('min_days', 3)
            
            employees = list(employees_collection.find())
            att_records = list(attendance_collection.find({'period': period}))
            
            att_data = {}
            for rec in att_records:
                emp_id = rec.get('employee_id')
                if emp_id and emp_id != '__init__':
                    if emp_id not in att_data:
                        att_data[emp_id] = {}
                    att_data[emp_id][rec.get('date')] = rec.get('status', '')
            
            absent_employees = []
            for emp in employees:
                emp_att = att_data.get(emp['_id'], {})
                absent_count = sum(1 for status in emp_att.values() if status == 'A')
                if absent_count >= min_days:
                    absent_employees.append({
                        'name': emp.get('name'),
                        'department': emp.get('department'),
                        'position': emp.get('position'),
                        'absent_days': absent_count
                    })
            
            if absent_employees:
                return {
                    "success": True,
                    "message": f"Employees with {min_days}+ absences:",
                    "data": absent_employees,
                    "type": "list"
                }
            else:
                return {
                    "success": True,
                    "message": f"No employees with {min_days}+ absences",
                    "data": [],
                    "type": "empty"
                }
        
        elif operation == 'get_best_attendance':
            period = params.get('period', get_current_period())
            limit = params.get('limit', 5)
            
            employees = list(employees_collection.find())
            att_records = list(attendance_collection.find({'period': period}))
            
            att_data = {}
            for rec in att_records:
                emp_id = rec.get('employee_id')
                if emp_id and emp_id != '__init__':
                    if emp_id not in att_data:
                        att_data[emp_id] = {}
                    att_data[emp_id][rec.get('date')] = rec.get('status', '')
            
            emp_stats = []
            for emp in employees:
                emp_att = att_data.get(emp['_id'], {})
                total = len(emp_att)
                if total > 0:
                    present = sum(1 for status in emp_att.values() if status == 'P')
                    rate = round((present / total * 100), 1)
                    emp_stats.append({
                        'name': emp.get('name'),
                        'department': emp.get('department'),
                        'position': emp.get('position'),
                        'attendance_rate': rate,
                        'present': present,
                        'total': total
                    })
            
            emp_stats.sort(key=lambda x: x['attendance_rate'], reverse=True)
            top_employees = emp_stats[:limit]
            
            return {
                "success": True,
                "message": f"Top {limit} employees by attendance:",
                "data": top_employees,
                "type": "list"
            }
        
        elif operation == 'get_worst_attendance':
            period = params.get('period', get_current_period())
            limit = params.get('limit', 5)
            
            employees = list(employees_collection.find())
            att_records = list(attendance_collection.find({'period': period}))
            
            att_data = {}
            for rec in att_records:
                emp_id = rec.get('employee_id')
                if emp_id and emp_id != '__init__':
                    if emp_id not in att_data:
                        att_data[emp_id] = {}
                    att_data[emp_id][rec.get('date')] = rec.get('status', '')
            
            emp_stats = []
            for emp in employees:
                emp_att = att_data.get(emp['_id'], {})
                total = len(emp_att)
                if total > 0:
                    present = sum(1 for status in emp_att.values() if status == 'P')
                    rate = round((present / total * 100), 1)
                    emp_stats.append({
                        'name': emp.get('name'),
                        'department': emp.get('department'),
                        'position': emp.get('position'),
                        'attendance_rate': rate,
                        'present': present,
                        'total': total
                    })
            
            emp_stats.sort(key=lambda x: x['attendance_rate'])
            bottom_employees = emp_stats[:limit]
            
            return {
                "success": True,
                "message": f"Bottom {limit} employees by attendance:",
                "data": bottom_employees,
                "type": "list"
            }
        
        elif operation == 'get_employees_by_day_off':
            day = params.get('day', '').strip()
            if not day:
                return {
                    "success": False,
                    "message": "Please specify a day (Monday, Tuesday, etc.)",
                    "type": "error"
                }
            
            employees = list(employees_collection.find({
                'day_off': {'$regex': day, '$options': 'i'}
            }))
            
            if employees:
                return {
                    "success": True,
                    "message": f"Employees with {day.title()} as day off:",
                    "data": employees,
                    "type": "list"
                }
            else:
                return {
                    "success": True,
                    "message": f"No employees with {day.title()} as day off",
                    "data": [],
                    "type": "empty"
                }
        
        elif operation == 'get_all_employees':
            employees = list(employees_collection.find())
            if employees:
                return {
                    "success": True,
                    "message": f"All employees ({len(employees)} total):",
                    "data": employees,
                    "type": "list"
                }
            else:
                return {
                    "success": True,
                    "message": "No employees found",
                    "data": [],
                    "type": "empty"
                }
        
        elif operation == 'get_employee_attendance_totals':
            name = params.get('employee_name', '').strip()
            if not name:
                return {
                    "success": False,
                    "message": "Please specify an employee name.",
                    "type": "error"
                }
            
            emp = employees_collection.find_one({
                'name': {'$regex': name, '$options': 'i'}
            })
            
            if not emp:
                return {
                    "success": True,
                    "message": f"No employee found matching '{name}'",
                    "data": [],
                    "type": "empty"
                }
            
            period = params.get('period', get_current_period())
            
            att_records = list(attendance_collection.find({
                'employee_id': emp['_id'],
                'period': period
            }))
            
            p = a = s = l = o = 0
            for rec in att_records:
                status = rec.get('status', '')
                if status == 'P': p += 1
                elif status == 'A': a += 1
                elif status == 'S': s += 1
                elif status == 'L': l += 1
                elif status == 'O': o += 1
            
            total = p + a + s + l + o
            rate = round((p / total * 100) if total > 0 else 0, 1)
            
            return {
                "success": True,
                "message": f"Attendance totals for {emp['name']} ({period}):",
                "data": {
                    'employee': emp['name'],
                    'department': emp.get('department'),
                    'position': emp.get('position'),
                    'period': period,
                    'present': p,
                    'absent': a,
                    'sick': s,
                    'leave': l,
                    'day_off': o,
                    'total_days': total,
                    'attendance_rate': rate
                },
                "type": "employee_stats"
            }
        
        else:
            return {
                "success": False,
                "message": "I couldn't find the information you're looking for.",
                "type": "error"
            }
            
    except Exception as e:
        print(f"Operation execution error: {e}")
        return {
            "success": False,
            "message": "Sorry, I encountered an error while retrieving data.",
            "type": "error"
        }

def get_current_period():
    now = datetime.datetime.now()
    day = now.day
    year = now.year
    month = now.month
    if day >= 27:
        month += 1
        if month > 12:
            month = 1
            year += 1
    return f"{year}-{str(month).zfill(2)}"

# ============================================================
# STATIC ROUTES
# ============================================================

@app.route('/')
def serve_login():
    return send_from_directory('public', 'login.html')

@app.route('/login')
def serve_login_page():
    return send_from_directory('public', 'login.html')

@app.route('/app')
def serve_app():
    return send_from_directory('public', 'index.html')

@app.route('/ai-assistant')
def serve_ai_assistant():
    return send_from_directory('public', 'ai-assistant.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('public', path)

@app.route('/health')
def health_check():
    return jsonify({
        "status": "healthy",
        "database": "connected" if client else "disconnected",
        "timestamp": datetime.datetime.now().isoformat()
    })

@app.route('/api/admin/attendance/rebuild-all', methods=['POST'])
@auth_required
@check_role(['IT Specialist'])
def rebuild_all_attendance():
    return jsonify({
        "success": True,
        "message": "This endpoint is not needed for the individual records structure"
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    
    if AUTO_MIGRATE:
        auto_migrate_data()
    else:
        print("Auto-migration disabled (set AUTO_MIGRATE=true to enable)")
    
    fix_existing_employee_numbers()
    initialize_default_users()
    ensure_period_exists()
    
    print("\n" + "=" * 55)
    print("    SILVER SANDS SALIMA HRMS")
    print("=" * 55)
    print(f"    Server running on port: {port}")
    print(f"    Database: MongoDB Atlas")
    print(f"    Mode: {'Development' if FLASK_DEBUG else 'Production'}")
    print(f"    Auto-Migration: {'ENABLED' if AUTO_MIGRATE else 'DISABLED'}")
    print(f"    OpenAI: {'ENABLED' if OPENAI_AVAILABLE else 'DISABLED'}")
    print("=" * 55)
    print(f"\nOpen: http://localhost:{port}")
    print("Login: developer / 192.168.1.1")
    print("Supervisor: supervisor / super2026")
    print("Admin: admin / smooth")
    print("Front Viewer: front / frontview")
    print("\nAI Assistant: http://localhost:{port}/ai-assistant")
    print("=" * 55 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=FLASK_DEBUG, threaded=True)